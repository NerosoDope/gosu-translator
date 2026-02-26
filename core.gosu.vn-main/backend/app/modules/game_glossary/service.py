"""
Game_Glossary Service - Business logic

Author: GOSU Development Team
Version: 1.0.0
"""

from typing import List, Optional, Dict, Any
from datetime import timezone, timedelta
from sqlalchemy.ext.asyncio import AsyncSession
from app.modules.game_glossary.repository import Game_GlossaryRepository
from app.modules.import_batches.service import ImportBatchService
from fastapi import UploadFile
import io
import logging

logger = logging.getLogger(__name__)


class Game_GlossaryService:
    """Game_Glossary Service - Business logic cho game_glossary"""
    
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = Game_GlossaryRepository(db)
        self.import_batch_svc = ImportBatchService(db)
    
    async def list(self, skip: int = 0, limit: int = 20, search: Optional[str] = None, is_active: Optional[bool] = None, language_pair: Optional[str] = None, game_id: Optional[int] = None, sort_by: Optional[str] = None, sort_order: Optional[str] = None) -> List[Dict[str, Any]]:
        """List game_glossary"""
        return await self.repo.list(skip=skip, limit=limit, search=search, is_active=is_active, language_pair=language_pair, game_id=game_id, sort_by=sort_by, sort_order=sort_order)
    
    async def find_translation(self, term: str, language_pair: str) -> Optional[str]:
        """Tìm bản dịch theo term và language_pair (exact match)."""
        return await self.repo.find_translation(term=term, language_pair=language_pair)

    async def get(self, id: int) -> Optional[Dict[str, Any]]:
        """Get game_glossary by ID"""
        return await self.repo.get(id)
    
    async def create(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Create game_glossary"""
        return await self.repo.create(data)
    
    async def update(self, id: int, data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Update game_glossary"""
        return await self.repo.update(id, data)
    
    async def delete(self, id: int) -> bool:
        """Delete game_glossary"""
        return await self.repo.delete(id)

    async def delete_all(self, game_id: Optional[int] = None) -> int:
        """Delete all game_glossary, optionally filter by game_id. Return count deleted."""
        return await self.repo.delete_all(game_id=game_id)
    
    async def upload_excel(self, file: UploadFile, game_id: Optional[int] = None) -> Dict[str, Any]:
        """
        Upload Excel file và import vào game_glossary
        
        Excel format:
        - Row 1: Header (term, translated_term, language_pair, game_id (optional), usage_count (optional), is_active (optional))
        - Row 2+: Data rows
        
        Args:
            file: Excel file to upload
            game_id: Game ID (nếu không có trong Excel, sẽ dùng giá trị này)
        
        Returns:
            Dict với thông tin kết quả upload
        """
        try:
            # Lazy import openpyxl để tránh lỗi khi module chưa được cài đặt
            try:
                import openpyxl
            except ImportError:
                return {
                    "success": False,
                    "total_rows": 0,
                    "created_count": 0,
                    "error_count": 1,
                    "errors": ["openpyxl module is not installed. Please install it: pip install openpyxl"]
                }
            
            # Đọc file Excel
            file_content = await file.read()
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            worksheet = workbook.active
            
            # Đọc header
            headers = [cell.value for cell in worksheet[1]]
            headers = [str(h).lower().strip() if h else "" for h in headers]
            
            # Chặn import chéo: file không có cột game_id = file từ điển chung, không được import vào từ điển game
            if "game_id" not in headers:
                return {
                    "success": False,
                    "total_rows": 0,
                    "created_count": 0,
                    "error_count": 1,
                    "errors": ["File Excel không có cột game_id (định dạng từ điển chung) không được phép import vào từ điển game. Vui lòng dùng file có cột game_id."]
                }

            # Validate headers
            required_headers = ['term', 'translated_term', 'language_pair']
            missing_headers = [h for h in required_headers if h not in headers]
            if missing_headers:
                return {
                    "success": False,
                    "total_rows": 0,
                    "created_count": 0,
                    "error_count": 0,
                    "errors": [f"Missing required headers: {', '.join(missing_headers)}"]
                }
            
            # Đọc data rows
            items = []
            errors = []
            total_rows = worksheet.max_row - 1  # Exclude header
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Tạo dict từ row
                    row_dict = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(row):
                            row_dict[header] = row[col_idx]
                    
                    # Validate và convert data
                    term = row_dict.get('term')
                    translated_term = row_dict.get('translated_term')
                    language_pair = row_dict.get('language_pair')
                    
                    if not term or not translated_term or not language_pair:
                        errors.append(f"Row {row_idx}: Missing required fields (term, translated_term, or language_pair)")
                        continue
                    
                    # Convert to string and strip
                    term = str(term).strip() if term else ""
                    translated_term = str(translated_term).strip() if translated_term else ""
                    language_pair = str(language_pair).strip() if language_pair else ""
                    
                    if not term or not translated_term or not language_pair:
                        errors.append(f"Row {row_idx}: Required fields cannot be empty")
                        continue
                    
                    # Lấy game_id từ Excel hoặc từ parameter
                    final_game_id = None
                    if row_dict.get('game_id') is not None:
                        try:
                            game_id_val = str(row_dict['game_id']).strip()
                            if game_id_val:
                                final_game_id = int(game_id_val)
                        except (ValueError, TypeError):
                            pass
                    
                    if not final_game_id and game_id:
                        final_game_id = game_id
                    
                    if not final_game_id:
                        errors.append(f"Row {row_idx}: Missing game_id")
                        continue
                    
                    # Parse optional fields
                    usage_count = 0
                    if row_dict.get('usage_count') is not None:
                        try:
                            usage_count_val = str(row_dict['usage_count']).strip()
                            if usage_count_val:
                                usage_count = int(usage_count_val)
                        except (ValueError, TypeError):
                            pass
                    
                    is_active = True
                    if row_dict.get('is_active') is not None:
                        is_active_val = row_dict['is_active']
                        if isinstance(is_active_val, bool):
                            is_active = is_active_val
                        elif isinstance(is_active_val, str):
                            is_active = is_active_val.lower().strip() in ('true', '1', 'yes', 'y')
                        elif isinstance(is_active_val, (int, float)):
                            is_active = bool(is_active_val)
                    
                    item_data = {
                        "term": term,
                        "translated_term": translated_term,
                        "language_pair": language_pair,
                        "game_id": final_game_id,
                        "usage_count": usage_count,
                        "is_active": is_active
                    }
                    
                    items.append(item_data)
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
                    logger.error(f"Error processing row {row_idx}: {e}")
            
            # Lọc bỏ các dòng đã tồn tại (term + translated_term + language_pair + game_id)
            existing_keys = await self.repo.get_existing_keys()
            items_to_create = []
            skipped_count = 0
            for item in items:
                key = (item["term"], item["translated_term"], item["language_pair"], item["game_id"])
                if key in existing_keys:
                    skipped_count += 1
                    continue
                items_to_create.append(item)
            
            # Tạo import batch và bulk create
            import_id = None
            created_count = 0
            batch_game_id = game_id or (items_to_create[0]["game_id"] if items_to_create else None)
            if items_to_create:
                import_id = await self.import_batch_svc.create(
                    source_type="game_glossary",
                    filename=file.filename or "unknown.xlsx",
                    total_rows=total_rows,
                    created_count=len(items_to_create),
                    error_count=len(errors),
                    game_id=batch_game_id,
                    user_id=None,
                )
                for item in items_to_create:
                    item["import_id"] = import_id
                created_items = await self.repo.bulk_create(items_to_create)
                created_count = len(created_items)
            
            return {
                "success": len(errors) == 0,
                "total_rows": total_rows,
                "created_count": created_count,
                "skipped_count": skipped_count,
                "error_count": len(errors),
                "errors": errors,
                "import_id": import_id,
            }
        except Exception as e:
            logger.error(f"Error uploading Excel file: {e}")
            return {
                "success": False,
                "total_rows": 0,
                "created_count": 0,
                "error_count": 1,
                "errors": [f"Error processing Excel file: {str(e)}"]
            }

    async def export_excel(self, game_id: Optional[int] = None) -> bytes:
        """
        Export Game Glossary ra file Excel.
        Nếu truyền game_id thì chỉ export glossary của game đó.
        Trả về bytes của file .xlsx.
        """
        # Lazy import openpyxl
        try:
            import openpyxl
        except ImportError as e:
            logger.error(f"openpyxl is not installed: {e}")
            raise

        # Lấy dữ liệu (dùng limit lớn để lấy nhiều bản ghi)
        items = await self.repo.list(
            skip=0,
            limit=100000,
            search=None,
            is_active=None,
            language_pair=None,
            game_id=game_id,
            sort_by=None,
            sort_order=None,
        )

        # Tạo workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Game Glossary"

        # Header
        headers = [
            "id",
            "import_id",
            "term",
            "translated_term",
            "language_pair",
            "game_id",
            "usage_count",
            "is_active",
            "created_at",
            "updated_at",
        ]
        worksheet.append(headers)

        GMT7 = timezone(timedelta(hours=7))

        def _excel_safe_value(v):
            """Format datetime for Excel - convert sang GMT+7."""
            if v is None:
                return None
            if hasattr(v, "strftime"):
                dt = v
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                dt_gmt7 = dt.astimezone(GMT7)
                return dt_gmt7.strftime("%Y-%m-%d %H:%M:%S")
            return v

        # Rows
        for item in items:
            worksheet.append([
                item.get("id"),
                item.get("import_id"),
                item.get("term"),
                item.get("translated_term"),
                item.get("language_pair"),
                item.get("game_id"),
                item.get("usage_count"),
                item.get("is_active"),
                _excel_safe_value(item.get("created_at")),
                _excel_safe_value(item.get("updated_at")),
            ])

        # Ghi ra bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()
