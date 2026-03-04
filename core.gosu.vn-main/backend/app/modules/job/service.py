"""
Job Service - Business logic

Author: GOSU Development Team
Version: 1.0.0
"""

import io
from typing import List, Optional, Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from app.modules.job.repository import JobRepository


class JobService:
    """Job Service - Business logic cho job"""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = JobRepository(db)

    async def list(
        self,
        skip: int = 0,
        limit: int = 20,
        query: Optional[str] = None,
        status: Optional[str] = None,
        job_type: Optional[str] = None,
        user_id: Optional[int] = None,
        include_deleted: bool = False,
        sort_by: str = "id",
        sort_order: str = "asc",
    ) -> Dict[str, Any]:
        items, total = await self.repo.list(
            skip=skip,
            limit=limit,
            query_str=query,
            status=status,
            job_type=job_type,
            user_id=user_id,
            include_deleted=include_deleted,
            sort_by=sort_by,
            sort_order=sort_order,
        )
        pages = (total + limit - 1) // limit if limit else 0
        page = (skip // limit) + 1 if limit else 1
        return {"items": items, "total": total, "page": page, "per_page": limit, "pages": pages}

    async def export_excel(
        self,
        query: Optional[str] = None,
        status: Optional[str] = None,
        job_type: Optional[str] = None,
        user_id: Optional[int] = None,
        include_deleted: bool = False,
    ) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise

        items = await self.repo.list_all(
            query_str=query, status=status, job_type=job_type,
            user_id=user_id, include_deleted=include_deleted, limit=100000
        )

        workbook = openpyxl.Workbook()

        # ── Màu theo trạng thái ──────────────────────────────────────────────
        STATUS_COLORS = {
            "pending":     "FFF3CD",
            "in_progress": "CCE5FF",
            "completed":   "D4EDDA",
            "failed":      "F8D7DA",
            "cancelled":   "E2E3E5",
        }
        HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
        HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
        SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
        SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        THIN   = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        def _fmt_dt(val) -> str:
            if not val:
                return ""
            try:
                from datetime import timezone
                if hasattr(val, "astimezone"):
                    val = val.astimezone(timezone.utc)
                return val.strftime("%d/%m/%Y %H:%M:%S")
            except Exception:
                return str(val)

        def _duration(started, finished) -> str:
            if not started or not finished:
                return ""
            try:
                diff = finished - started
                total = int(diff.total_seconds())
                if total < 0:
                    return ""
                h, rem = divmod(total, 3600)
                m, s   = divmod(rem, 60)
                if h:
                    return f"{h}g {m}p {s}s"
                if m:
                    return f"{m}p {s}s"
                return f"{s}s"
            except Exception:
                return ""

        def _style_header_row(ws, row: int, fill, font, col_count: int):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill   = fill
                cell.font   = font
                cell.alignment = CENTER
                cell.border = THIN

        def _style_data_row(ws, row: int, col_count: int, bg_hex: str = None):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                if bg_hex:
                    cell.fill = PatternFill("solid", fgColor=bg_hex)
                cell.alignment = LEFT if col > 2 else CENTER
                cell.border = THIN

        # ════════════════════════════════════════════════════════════════════
        # SHEET 1 — Tổng quan Jobs
        # ════════════════════════════════════════════════════════════════════
        ws1 = workbook.active
        ws1.title = "Tổng quan Jobs"

        headers1 = [
            "STT", "Mã Job", "Loại Job", "Trạng thái", "Tiến độ (%)",
            "Người tạo", "Ngôn ngữ nguồn", "Ngôn ngữ đích",
            "Độ ưu tiên", "Số lần thử lại", "Giới hạn retry",
            "Thời gian xử lý", "Ngày tạo", "Bắt đầu lúc", "Hoàn thành lúc",
            "Cập nhật lúc", "Thông báo lỗi", "Đã xóa",
        ]
        col_widths1 = [6, 28, 14, 14, 12, 20, 16, 16, 12, 14, 14, 16, 20, 20, 20, 20, 40, 10]

        ws1.append(headers1)
        _style_header_row(ws1, 1, HEADER_FILL, HEADER_FONT, len(headers1))

        for idx, item in enumerate(items, start=1):
            status_val = item.get("status") or ""
            started    = item.get("started_at")
            finished   = item.get("finished_at")
            row_data = [
                idx,
                item.get("job_code"),
                item.get("job_type"),
                status_val,
                item.get("progress") if item.get("progress") is not None else 0,
                item.get("creator_name") or str(item.get("user_id") or ""),
                item.get("source_lang"),
                item.get("target_lang"),
                item.get("priority"),
                item.get("retry_count") if item.get("retry_count") is not None else 0,
                item.get("max_retry")   if item.get("max_retry")   is not None else 3,
                _duration(started, finished),
                _fmt_dt(item.get("created_at")),
                _fmt_dt(started),
                _fmt_dt(finished),
                _fmt_dt(item.get("updated_at")),
                item.get("error_message") or "",
                "Có" if item.get("is_deleted") else "",
            ]
            ws1.append(row_data)
            row_num = idx + 1
            bg = STATUS_COLORS.get(status_val)
            _style_data_row(ws1, row_num, len(headers1), bg)

        for i, width in enumerate(col_widths1, start=1):
            ws1.column_dimensions[get_column_letter(i)].width = width
        ws1.row_dimensions[1].height = 30
        ws1.freeze_panes = "A2"

        # ════════════════════════════════════════════════════════════════════
        # SHEET 2 — Nội dung dịch
        # ════════════════════════════════════════════════════════════════════
        ws2 = workbook.create_sheet("Nội dung dịch")

        headers2 = [
            "STT", "Mã Job", "Trạng thái",
            "Ngôn ngữ nguồn", "Ngôn ngữ đích",
            "Văn bản gốc", "Kết quả dịch",
            "Ngày hoàn thành",
        ]
        col_widths2 = [6, 28, 14, 16, 16, 60, 60, 20]

        ws2.append(headers2)
        _style_header_row(ws2, 1, SUBHEADER_FILL, SUBHEADER_FONT, len(headers2))

        translation_items = [
            it for it in items
            if (it.get("payload") and it["payload"].get("text")) or
               (it.get("result")  and it["result"].get("translated"))
        ]

        for idx, item in enumerate(translation_items, start=1):
            payload    = item.get("payload") or {}
            result     = item.get("result")  or {}
            status_val = item.get("status") or ""
            row_data = [
                idx,
                item.get("job_code"),
                status_val,
                item.get("source_lang"),
                item.get("target_lang"),
                payload.get("text", ""),
                result.get("translated", ""),
                _fmt_dt(item.get("finished_at")),
            ]
            ws2.append(row_data)
            row_num = idx + 1
            bg = STATUS_COLORS.get(status_val)
            _style_data_row(ws2, row_num, len(headers2), bg)

        for i, width in enumerate(col_widths2, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = width
        ws2.row_dimensions[1].height = 30
        ws2.freeze_panes = "A2"

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    async def get(self, id: int, include_deleted: bool = False) -> Optional[Dict[str, Any]]:
        return await self.repo.get(id, include_deleted=include_deleted)

    async def create(self, data: Dict[str, Any]) -> Dict[str, Any]:
        return await self.repo.create(data)

    async def update(self, id: int, data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        return await self.repo.update(id, data)

    async def delete(self, id: int) -> bool:
        """Soft delete."""
        return await self.repo.delete(id)

    async def restore(self, id: int) -> Optional[Dict[str, Any]]:
        """Khôi phục job đã soft-delete."""
        return await self.repo.restore(id)

    async def hard_delete(self, id: int) -> bool:
        """Xóa vĩnh viễn."""
        return await self.repo.hard_delete(id)

    async def cancel(self, id: int) -> Optional[Dict[str, Any]]:
        """Hủy job đang pending/in_progress."""
        return await self.repo.cancel(id)

    async def retry(self, id: int) -> Optional[Dict[str, Any]]:
        """Thử lại job failed/cancelled."""
        return await self.repo.retry(id)
