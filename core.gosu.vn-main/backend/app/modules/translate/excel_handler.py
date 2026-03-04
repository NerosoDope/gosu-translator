"""
Xử lý file Excel (.xlsx, .xls): đọc, parse preview, đọc full, xuất.
"""
import io
from typing import List, Tuple

from fastapi import HTTPException

from app.modules.translate.schemas import ParseFileResponse


def cell_to_str(value) -> str:
    """Chuyển giá trị ô (Excel/CSV) sang chuỗi."""
    if value is None:
        return ""
    if hasattr(value, "isoformat"):  # datetime.date/datetime
        return value.isoformat()
    return str(value).strip()


async def parse_excel(content: bytes, preview_limit: int = 5) -> ParseFileResponse:
    """Đọc file .xlsx (openpyxl), trả về columns và preview_rows.
    Bỏ qua cột trống trong header và hàng mà toàn bộ cell đều trống.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=503,
            detail="openpyxl chưa được cài đặt. Vui lòng cài: pip install openpyxl",
        )
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    if not sheet:
        return ParseFileResponse(columns=[], preview_rows=[])
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return ParseFileResponse(columns=[], preview_rows=[])

    raw_columns = [cell_to_str(h) for h in header_row]
    valid_col_indices = [i for i, c in enumerate(raw_columns) if c.strip()]

    seen: dict = {}
    columns: List[str] = []
    for i in valid_col_indices:
        c = raw_columns[i]
        key = c
        if key in seen:
            seen[key] += 1
            c = f"{c}_{seen[key]}"
        else:
            seen[key] = 1
        columns.append(c)

    preview_rows = []
    for row in sheet.iter_rows(min_row=2, max_row=1 + preview_limit, values_only=True):
        row_dict = {
            col_name: (cell_to_str(row[orig_i]) if orig_i < len(row) else "")
            for col_name, orig_i in zip(columns, valid_col_indices)
        }
        if not any(v.strip() for v in row_dict.values()):
            continue
        preview_rows.append(row_dict)
    workbook.close()
    return ParseFileResponse(columns=columns, preview_rows=preview_rows)


def read_full_excel(content: bytes, max_rows: int) -> Tuple[List[str], List[dict]]:
    """Đọc toàn bộ file .xlsx (openpyxl), trả về columns và toàn bộ rows (tối đa max_rows dòng dữ liệu).
    Bỏ qua cột trống trong header và hàng mà toàn bộ cell đều trống.
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=503,
            detail="openpyxl chưa được cài đặt. Vui lòng cài: pip install openpyxl",
        )
    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    if not sheet:
        return [], []
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        workbook.close()
        return [], []

    raw_columns = [cell_to_str(h) for h in header_row]
    valid_col_indices = [i for i, c in enumerate(raw_columns) if c.strip()]

    seen: dict = {}
    columns: List[str] = []
    for i in valid_col_indices:
        c = raw_columns[i]
        key = c
        if key in seen:
            seen[key] += 1
            c = f"{c}_{seen[key]}"
        else:
            seen[key] = 1
        columns.append(c)

    rows = []
    counted = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if counted >= max_rows:
            break
        row_dict = {
            col_name: (cell_to_str(row[orig_i]) if orig_i < len(row) else "")
            for col_name, orig_i in zip(columns, valid_col_indices)
        }
        if not any(v.strip() for v in row_dict.values()):
            continue
        rows.append(row_dict)
        counted += 1
    workbook.close()
    return columns, rows


def export_xlsx(columns: List[str], rows: List[dict]) -> bytes:
    """Xuất columns + rows ra file .xlsx."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=503, detail="openpyxl chưa được cài đặt.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(columns)
    for row in rows:
        ws.append([str(row.get(c, "") or "") for c in columns])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
